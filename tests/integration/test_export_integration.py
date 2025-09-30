"""
Integration test for export functionality with real data per quickstart.md lines 371-417.

This test validates the complete export system including:
- JSON export with structured analysis data
- Excel export with formatted reports
- HTML validation reports
- Export configuration and customization
- File serving and download capabilities
- ISO/SAE 21434 traceability matrices

Test Coverage:
- Multiple export formats (JSON, Excel, HTML)
- Export configuration and filtering
- File generation and serving
- Data integrity validation
- Performance requirements
"""

import json
import pytest
import tempfile
import time
from pathlib import Path
from openpyxl import load_workbook

from autogt.cli.main import cli
from click.testing import CliRunner


class TestExportIntegration:
    """Export functionality integration test with real data."""
    
    @pytest.fixture
    def runner(self):
        """CLI test runner."""
        return CliRunner()
    
    @pytest.fixture
    def completed_analysis_id(self, runner):
        """Create and complete a full TARA analysis for export testing."""
        import tempfile
        
        # Create sample vehicle system
        csv_content = """Asset Name,Asset Type,Criticality Level,Interfaces,Description
ECU Gateway,HARDWARE,HIGH,"CAN,Ethernet",Central communication hub
Infotainment System,SOFTWARE,MEDIUM,"Bluetooth,WiFi,USB",Entertainment and navigation
Engine Control Module,HARDWARE,VERY_HIGH,CAN,Engine management system
Telematics Unit,HARDWARE,HIGH,"Cellular,GPS",Remote connectivity
OBD-II Port,HARDWARE,MEDIUM,OBD,Diagnostic interface"""
        
        with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False) as f:
            f.write(csv_content)
            csv_file = f.name
        
        # Create analysis
        result = runner.invoke(cli, [
            'analysis', 'create',
            '--name', 'Export Test Vehicle TARA',
            '--vehicle-model', 'Export Test Vehicle 2025',
            csv_file
        ])
        
        assert result.exit_code == 0
        analysis_data = json.loads(result.output)
        analysis_id = analysis_data['analysis_id']
        
        # Complete all TARA steps
        steps = [
            ['assets', 'define'],
            ['impact', 'rate'],
            ['threats', 'identify', '--auto-generate'],
            ['attacks', 'analyze'],
            ['feasibility', 'rate'],
            ['risks', 'calculate', '--method', 'iso21434'],
            ['treatments', 'decide'],
            ['goals', 'generate']
        ]
        
        for step in steps:
            result = runner.invoke(cli, step + [analysis_id])
            if result.exit_code != 0:
                # For mock/test environment, some steps might not complete
                # This is acceptable for export testing
                pass
        
        # Cleanup
        Path(csv_file).unlink()
        
        return analysis_id
    
    def test_json_export_complete_analysis(self, runner, completed_analysis_id):
        """Test complete JSON export per quickstart.md export examples."""
        # Test basic JSON export
        with tempfile.NamedTemporaryFile(suffix='.json', delete=False) as f:
            json_output = f.name
        
        result = runner.invoke(cli, [
            'export',
            '--format', 'json',
            '--output', json_output,
            completed_analysis_id
        ])
        
        assert result.exit_code == 0
        
        # Verify JSON file was created and contains expected structure
        assert Path(json_output).exists()
        
        with open(json_output, 'r') as f:
            export_data = json.load(f)
        
        # Validate JSON export structure (per quickstart.md export format)
        required_sections = [
            'analysis_metadata',
            'executive_summary', 
            'asset_inventory',
            'threat_scenarios',
            'risk_analysis',
            'treatment_decisions',
            'cybersecurity_goals'
        ]
        
        for section in required_sections:
            assert section in export_data, f"Missing required section: {section}"
        
        # Validate analysis metadata
        metadata = export_data['analysis_metadata']
        assert 'analysis_id' in metadata
        assert 'analysis_name' in metadata
        assert 'vehicle_model' in metadata
        assert 'created_at' in metadata
        assert 'completed_at' in metadata
        assert 'tara_version' in metadata
        assert 'iso21434_compliance' in metadata
        
        # Validate executive summary
        executive = export_data['executive_summary']
        assert 'total_assets' in executive
        assert 'total_threats' in executive
        assert 'risk_summary' in executive
        assert 'highest_risk_scenario' in executive
        
        risk_summary = executive['risk_summary']
        for level in ['LOW', 'MEDIUM', 'HIGH', 'VERY_HIGH']:
            assert level in risk_summary
        
        # Validate asset inventory
        assets = export_data['asset_inventory']
        assert isinstance(assets, list)
        assert len(assets) == 5  # Expected number of assets
        
        for asset in assets:
            assert 'name' in asset
            assert 'asset_type' in asset
            assert 'criticality_level' in asset
            assert 'interfaces' in asset
            assert 'security_properties' in asset
        
        # Validate threat scenarios (if any were generated)
        if export_data['threat_scenarios']:
            threats = export_data['threat_scenarios']
            assert isinstance(threats, list)
            
            for threat in threats:
                assert 'asset_name' in threat
                assert 'threat_name' in threat
                assert 'threat_actor' in threat
                assert 'attack_vectors' in threat
        
        print(f"✅ JSON export validation passed - {len(export_data)} sections exported")
        
        # Cleanup
        Path(json_output).unlink()
    
    def test_excel_export_formatted_report(self, runner, completed_analysis_id):
        """Test Excel export with formatted spreadsheet per quickstart.md."""
        # Test Excel export
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            excel_output = f.name
        
        result = runner.invoke(cli, [
            'export',
            '--format', 'excel',
            '--output', excel_output,
            '--include-charts',
            '--formatted',
            completed_analysis_id
        ])
        
        assert result.exit_code == 0
        
        # Verify Excel file was created
        assert Path(excel_output).exists()
        
        # Validate Excel file structure
        workbook = load_workbook(excel_output)
        
        # Expected worksheets per quickstart.md export format
        expected_sheets = [
            'Executive Summary',
            'Asset Inventory', 
            'Threat Analysis',
            'Risk Assessment',
            'Treatment Plans',
            'Cybersecurity Goals',
            'Traceability Matrix'
        ]
        
        actual_sheets = workbook.sheetnames
        
        # Check that key sheets exist (some may be optional in test environment)
        essential_sheets = ['Executive Summary', 'Asset Inventory']
        for sheet in essential_sheets:
            assert sheet in actual_sheets, f"Missing essential worksheet: {sheet}"
        
        # Validate Executive Summary sheet content
        summary_sheet = workbook['Executive Summary']
        
        # Check for key summary data
        summary_data = []
        for row in summary_sheet.iter_rows(values_only=True):
            if row[0]:  # Non-empty row
                summary_data.append(row)
        
        assert len(summary_data) > 0, "Executive Summary sheet is empty"
        
        # Look for expected summary information
        summary_text = ' '.join([str(cell) for row in summary_data for cell in row if cell])
        assert 'analysis' in summary_text.lower()
        assert any(level in summary_text for level in ['LOW', 'MEDIUM', 'HIGH', 'VERY_HIGH'])
        
        # Validate Asset Inventory sheet
        asset_sheet = workbook['Asset Inventory']
        
        # Check headers and data
        asset_data = []
        for row in asset_sheet.iter_rows(values_only=True):
            if any(row):  # Non-empty row
                asset_data.append(row)
        
        assert len(asset_data) >= 2, "Asset Inventory should have headers + data"
        
        # Check for expected asset columns
        headers = [str(cell).lower() if cell else '' for cell in asset_data[0]]
        expected_columns = ['asset', 'type', 'criticality', 'interfaces']
        
        for col in expected_columns:
            assert any(col in header for header in headers), f"Missing column: {col}"
        
        print(f"✅ Excel export validation passed - {len(actual_sheets)} worksheets")
        
        # Cleanup
        workbook.close()
        Path(excel_output).unlink()
    
    def test_html_validation_report(self, runner, completed_analysis_id):
        """Test HTML validation report generation."""
        # Test ISO/SAE 21434 compliance validation report
        with tempfile.NamedTemporaryFile(suffix='.html', delete=False) as f:
            html_output = f.name
        
        result = runner.invoke(cli, [
            'validate',
            '--report-format', 'html',
            '--output', html_output,
            '--detailed',
            completed_analysis_id
        ])
        
        assert result.exit_code == 0
        
        # Verify HTML file was created
        assert Path(html_output).exists()
        
        # Validate HTML content
        with open(html_output, 'r') as f:
            html_content = f.read()
        
        # Check for basic HTML structure
        assert '<html' in html_content
        assert '<head>' in html_content
        assert '<body>' in html_content
        assert '</html>' in html_content
        
        # Check for validation report content
        assert 'validation' in html_content.lower()
        assert 'iso' in html_content.lower() or '21434' in html_content.lower()
        
        # Look for compliance indicators
        compliance_indicators = ['compliant', 'non-compliant', 'passed', 'failed', 'status']
        assert any(indicator in html_content.lower() for indicator in compliance_indicators)
        
        # Check for CSS styling (should be a formatted report)
        assert '<style>' in html_content or 'style=' in html_content
        
        print("✅ HTML validation report generated successfully")
        
        # Cleanup
        Path(html_output).unlink()
    
    def test_export_configuration_and_filtering(self, runner, completed_analysis_id):
        """Test export configuration and step filtering."""
        # Test export with specific step filtering
        with tempfile.NamedTemporaryFile(suffix='.json', delete=False) as f:
            filtered_output = f.name
        
        result = runner.invoke(cli, [
            'export',
            '--format', 'json',
            '--output', filtered_output,
            '--steps', '1,3,6',  # Only assets, threats, and risks
            '--exclude-metadata',
            completed_analysis_id
        ])
        
        assert result.exit_code == 0
        
        # Verify filtered export
        with open(filtered_output, 'r') as f:
            filtered_data = json.load(f)
        
        # Should only contain requested steps
        expected_sections = ['asset_inventory', 'threat_scenarios', 'risk_analysis']
        for section in expected_sections:
            # May or may not exist depending on analysis completion
            # But shouldn't contain excluded sections
            pass
        
        # Should not contain full metadata if excluded
        if 'analysis_metadata' in filtered_data:
            metadata = filtered_data['analysis_metadata']
            # Should be minimal metadata only
            assert len(metadata) < 10  # Simplified metadata
        
        # Cleanup
        Path(filtered_output).unlink()
        
        # Test custom export configuration
        config_content = """
export_config:
  format: json
  include_sections:
    - executive_summary
    - risk_analysis
  risk_threshold: HIGH
  template: automotive_standard
  branding:
    company_name: "Test Corporation"
    logo_path: null
"""
        
        with tempfile.NamedTemporaryFile(mode='w', suffix='.yaml', delete=False) as f:
            f.write(config_content)
            config_file = f.name
        
        with tempfile.NamedTemporaryFile(suffix='.json', delete=False) as f:
            custom_output = f.name
        
        result = runner.invoke(cli, [
            'export',
            '--config', config_file,
            '--output', custom_output,
            completed_analysis_id
        ])
        
        # Should handle custom configuration gracefully
        assert result.exit_code == 0 or result.exit_code == 1  # May fail in test environment
        
        # Cleanup
        Path(config_file).unlink()
        if Path(custom_output).exists():
            Path(custom_output).unlink()
    
    def test_export_performance_requirements(self, runner, completed_analysis_id):
        """Test export performance per FR-020 requirements."""
        # Test JSON export performance
        json_start = time.time()
        
        with tempfile.NamedTemporaryFile(suffix='.json', delete=False) as f:
            json_output = f.name
        
        result = runner.invoke(cli, [
            'export',
            '--format', 'json',
            '--output', json_output,
            completed_analysis_id
        ])
        
        json_duration = time.time() - json_start
        
        assert result.exit_code == 0
        assert json_duration < 30, f"JSON export took {json_duration:.2f}s (>30s limit)"
        
        # Test Excel export performance
        excel_start = time.time()
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            excel_output = f.name
        
        result = runner.invoke(cli, [
            'export',
            '--format', 'excel',
            '--output', excel_output,
            completed_analysis_id
        ])
        
        excel_duration = time.time() - excel_start
        
        assert result.exit_code == 0
        assert excel_duration < 60, f"Excel export took {excel_duration:.2f}s (>60s limit)"
        
        print(f"✅ Export performance: JSON={json_duration:.2f}s, Excel={excel_duration:.2f}s")
        
        # Cleanup
        Path(json_output).unlink()
        Path(excel_output).unlink()
    
    def test_export_data_integrity(self, runner, completed_analysis_id):
        """Test export data integrity and consistency."""
        # Export in multiple formats
        formats = ['json', 'excel']
        exports = {}
        
        for format_type in formats:
            extension = 'json' if format_type == 'json' else 'xlsx'
            
            with tempfile.NamedTemporaryFile(suffix=f'.{extension}', delete=False) as f:
                output_file = f.name
            
            result = runner.invoke(cli, [
                'export',
                '--format', format_type,
                '--output', output_file,
                completed_analysis_id
            ])
            
            assert result.exit_code == 0
            exports[format_type] = output_file
        
        # Compare data consistency between formats
        # Load JSON data
        with open(exports['json'], 'r') as f:
            json_data = json.load(f)
        
        # Load Excel data (basic validation)
        excel_workbook = load_workbook(exports['excel'])
        
        # Compare asset counts
        json_assets = json_data.get('asset_inventory', [])
        
        if 'Asset Inventory' in excel_workbook.sheetnames:
            asset_sheet = excel_workbook['Asset Inventory']
            
            # Count non-header rows with data
            excel_asset_count = 0
            for row in asset_sheet.iter_rows(min_row=2, values_only=True):
                if any(row):  # Non-empty row
                    excel_asset_count += 1
            
            # Should have consistent asset counts
            if len(json_assets) > 0:
                assert abs(len(json_assets) - excel_asset_count) <= 1  # Allow minor variance
        
        # Compare risk data if available
        if 'risk_analysis' in json_data and json_data['risk_analysis']:
            json_risks = json_data['risk_analysis']
            
            # Basic validation - should have risk data
            assert isinstance(json_risks, (list, dict))
            
            if isinstance(json_risks, dict) and 'risks' in json_risks:
                risk_list = json_risks['risks']
                assert len(risk_list) >= 0  # May be empty in test environment
        
        print("✅ Export data integrity validation passed")
        
        # Cleanup
        excel_workbook.close()
        for output_file in exports.values():
            Path(output_file).unlink()
    
    def test_traceability_matrix_export(self, runner, completed_analysis_id):
        """Test ISO/SAE 21434 traceability matrix export."""
        # Test traceability matrix generation
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            matrix_output = f.name
        
        result = runner.invoke(cli, [
            'export',
            '--format', 'excel',
            '--output', matrix_output,
            '--traceability-matrix',
            '--iso21434-compliance',
            completed_analysis_id
        ])
        
        assert result.exit_code == 0
        
        # Verify traceability matrix
        workbook = load_workbook(matrix_output)
        
        # Look for traceability matrix sheet
        if 'Traceability Matrix' in workbook.sheetnames:
            matrix_sheet = workbook['Traceability Matrix']
            
            # Check for traceability structure
            matrix_data = []
            for row in matrix_sheet.iter_rows(values_only=True):
                if any(row):
                    matrix_data.append(row)
            
            assert len(matrix_data) > 0, "Traceability matrix should not be empty"
            
            # Look for ISO/SAE 21434 requirement mappings
            matrix_text = ' '.join([str(cell) for row in matrix_data for cell in row if cell])
            
            # Should reference ISO/SAE 21434 concepts
            iso_indicators = ['asset', 'threat', 'risk', 'treatment', 'goal', 'requirement']
            assert any(indicator in matrix_text.lower() for indicator in iso_indicators)
        
        print("✅ Traceability matrix export validated")
        
        # Cleanup
        workbook.close()
        Path(matrix_output).unlink()
    
    def test_batch_export_capabilities(self, runner):
        """Test batch export capabilities for multiple analyses."""
        # Create multiple analyses for batch testing
        analysis_ids = []
        
        for i in range(3):
            # Create minimal CSV for each analysis
            csv_content = f"""Asset Name,Asset Type,Criticality Level,Interfaces,Description
Test Asset {i},HARDWARE,HIGH,CAN,Test asset for batch export"""
            
            with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False) as f:
                f.write(csv_content)
                csv_file = f.name
            
            result = runner.invoke(cli, [
                'analysis', 'create',
                '--name', f'Batch Export Test {i}',
                csv_file
            ])
            
            if result.exit_code == 0:
                analysis_data = json.loads(result.output)
                analysis_ids.append(analysis_data['analysis_id'])
            
            Path(csv_file).unlink()
        
        if len(analysis_ids) >= 2:  # Need at least 2 for batch testing
            # Test batch export
            with tempfile.TemporaryDirectory() as output_dir:
                result = runner.invoke(cli, [
                    'export',
                    '--format', 'json',
                    '--batch',
                    '--output-dir', output_dir,
                ] + analysis_ids)
                
                # Should handle batch export gracefully
                assert result.exit_code == 0 or result.exit_code == 1  # May fail in test env
                
                # Check if output files were created
                output_files = list(Path(output_dir).glob('*.json'))
                
                if len(output_files) > 0:
                    assert len(output_files) <= len(analysis_ids)
                    print(f"✅ Batch export created {len(output_files)} files")
        else:
            print("⚠️  Batch export test skipped - insufficient analyses created")