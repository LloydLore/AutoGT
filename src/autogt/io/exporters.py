"""Export handlers for TARA analysis results to various formats.

Provides structured export capabilities for analysis results including
Excel, JSON, CSV, and PDF formats with automotive compliance features.
"""

import json
import csv
from typing import Dict, List, Any, Optional, Union
from pathlib import Path
import logging
from datetime import datetime
from dataclasses import dataclass
import pandas as pd

from ..core.exceptions import ExportError, ValidationError


logger = logging.getLogger(__name__)


@dataclass
class ExportResult:
    """Result of export operation with metadata."""
    success: bool
    output_path: Path
    format_used: str
    file_size_bytes: int
    export_time_ms: int
    items_exported: Dict[str, int]
    metadata: Dict[str, Any]
    warnings: List[str]


class BaseExporter:
    """Base class for all export handlers."""
    
    def __init__(self, format_name: str):
        self.format_name = format_name
        self.export_timestamp = datetime.now()
    
    def export(self, data: Dict[str, Any], output_path: Path, **options) -> ExportResult:
        """Export data to specified format."""
        raise NotImplementedError("Subclasses must implement export method")
    
    def _validate_data(self, data: Dict[str, Any]) -> List[str]:
        """Validate data before export."""
        errors = []
        
        if not isinstance(data, dict):
            errors.append("Data must be a dictionary")
            return errors
        
        # Check for required sections
        if 'assets' not in data and 'threats' not in data:
            errors.append("Data must contain 'assets' or 'threats' section")
        
        # Validate assets structure
        if 'assets' in data:
            if not isinstance(data['assets'], list):
                errors.append("Assets must be a list")
            else:
                for i, asset in enumerate(data['assets']):
                    if not isinstance(asset, dict):
                        errors.append(f"Asset {i} must be a dictionary")
                    elif 'name' not in asset:
                        errors.append(f"Asset {i} missing required 'name' field")
        
        # Validate threats structure
        if 'threats' in data:
            if not isinstance(data['threats'], list):
                errors.append("Threats must be a list")
            else:
                for i, threat in enumerate(data['threats']):
                    if not isinstance(threat, dict):
                        errors.append(f"Threat {i} must be a dictionary")
                    elif 'name' not in threat:
                        errors.append(f"Threat {i} missing required 'name' field")
        
        return errors
    
    def _get_export_metadata(self, data: Dict[str, Any]) -> Dict[str, Any]:
        """Generate export metadata."""
        return {
            'export_timestamp': self.export_timestamp.isoformat(),
            'exporter': self.format_name,
            'asset_count': len(data.get('assets', [])),
            'threat_count': len(data.get('threats', [])),
            'source_metadata': data.get('metadata', {}),
            'export_version': '1.0.0'
        }


class JsonExporter(BaseExporter):
    """Export TARA analysis results to JSON format."""
    
    def __init__(self):
        super().__init__('json')
    
    def export(self, data: Dict[str, Any], output_path: Path, **options) -> ExportResult:
        """Export data to JSON format.
        
        Args:
            data: Dictionary containing assets, threats, and metadata
            output_path: Path for output JSON file
            **options: Additional options (indent, ensure_ascii, etc.)
        
        Returns:
            ExportResult with export status and metadata
        """
        start_time = datetime.now()
        warnings = []
        
        try:
            # Validate input data
            validation_errors = self._validate_data(data)
            if validation_errors:
                raise ValidationError(f"Data validation failed: {'; '.join(validation_errors)}")
            
            # Prepare export data structure
            export_data = self._prepare_json_structure(data, **options)
            
            # Configure JSON export options
            json_options = {
                'indent': options.get('indent', 2),
                'ensure_ascii': options.get('ensure_ascii', False),
                'sort_keys': options.get('sort_keys', True)
            }
            
            # Write JSON file
            output_path.parent.mkdir(parents=True, exist_ok=True)
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(export_data, f, **json_options)
            
            # Calculate export metrics
            processing_time = int((datetime.now() - start_time).total_seconds() * 1000)
            file_size = output_path.stat().st_size
            
            items_exported = {
                'assets': len(data.get('assets', [])),
                'threats': len(data.get('threats', [])),
                'total_items': len(data.get('assets', [])) + len(data.get('threats', []))
            }
            
            return ExportResult(
                success=True,
                output_path=output_path,
                format_used='json',
                file_size_bytes=file_size,
                export_time_ms=processing_time,
                items_exported=items_exported,
                metadata=self._get_export_metadata(data),
                warnings=warnings
            )
            
        except Exception as e:
            logger.error(f"JSON export failed: {str(e)}")
            raise ExportError(f"JSON export failed: {str(e)}", format='json')
    
    def _prepare_json_structure(self, data: Dict[str, Any], **options) -> Dict[str, Any]:
        """Prepare JSON structure for export."""
        export_structure = {
            'metadata': self._get_export_metadata(data),
            'assets': data.get('assets', []),
            'threats': data.get('threats', [])
        }
        
        # Add analysis metadata if present
        if 'analysis' in data:
            export_structure['analysis'] = data['analysis']
        
        # Add confidence and validation info if present
        if 'confidence_score' in data:
            export_structure['confidence_score'] = data['confidence_score']
        
        if 'validation_errors' in data:
            export_structure['validation_errors'] = data['validation_errors']
        
        # Include source file information if available
        if 'source_files' in data:
            export_structure['source_files'] = data['source_files']
        
        return export_structure


class CsvExporter(BaseExporter):
    """Export TARA analysis results to CSV format."""
    
    def __init__(self):
        super().__init__('csv')
    
    def export(self, data: Dict[str, Any], output_path: Path, **options) -> ExportResult:
        """Export data to CSV format.
        
        Args:
            data: Dictionary containing assets, threats, and metadata
            output_path: Path for output CSV file (or directory for multiple files)
            **options: Additional options (separate_files, delimiter, etc.)
        
        Returns:
            ExportResult with export status and metadata
        """
        start_time = datetime.now()
        warnings = []
        
        try:
            # Validate input data
            validation_errors = self._validate_data(data)
            if validation_errors:
                raise ValidationError(f"Data validation failed: {'; '.join(validation_errors)}")
            
            # Determine export strategy
            separate_files = options.get('separate_files', True)
            
            if separate_files:
                result = self._export_separate_csv_files(data, output_path, **options)
            else:
                result = self._export_single_csv_file(data, output_path, **options)
            
            # Update timing
            processing_time = int((datetime.now() - start_time).total_seconds() * 1000)
            result.export_time_ms = processing_time
            
            return result
            
        except Exception as e:
            logger.error(f"CSV export failed: {str(e)}")
            raise ExportError(f"CSV export failed: {str(e)}", format='csv')
    
    def _export_separate_csv_files(self, data: Dict[str, Any], 
                                  output_path: Path, **options) -> ExportResult:
        """Export assets and threats to separate CSV files."""
        output_path.mkdir(parents=True, exist_ok=True)
        
        csv_options = {
            'delimiter': options.get('delimiter', ','),
            'quoting': options.get('quoting', csv.QUOTE_MINIMAL),
            'lineterminator': '\n'
        }
        
        total_size = 0
        items_exported = {'assets': 0, 'threats': 0}
        
        # Export assets
        if 'assets' in data and data['assets']:
            assets_file = output_path / 'assets.csv'
            assets_df = pd.DataFrame(data['assets'])
            
            # Ensure consistent column order
            asset_columns = ['name', 'asset_type', 'criticality', 'interfaces', 
                           'description', 'location', 'manufacturer']
            assets_df = self._reorder_columns(assets_df, asset_columns)
            
            assets_df.to_csv(assets_file, index=False, **csv_options)
            total_size += assets_file.stat().st_size
            items_exported['assets'] = len(data['assets'])
        
        # Export threats
        if 'threats' in data and data['threats']:
            threats_file = output_path / 'threats.csv'
            threats_df = pd.DataFrame(data['threats'])
            
            # Ensure consistent column order
            threat_columns = ['name', 'category', 'description', 'likelihood', 
                            'impact', 'target_asset']
            threats_df = self._reorder_columns(threats_df, threat_columns)
            
            threats_df.to_csv(threats_file, index=False, **csv_options)
            total_size += threats_file.stat().st_size
            items_exported['threats'] = len(data['threats'])
        
        # Export metadata
        metadata_file = output_path / 'metadata.json'
        with open(metadata_file, 'w', encoding='utf-8') as f:
            json.dump(self._get_export_metadata(data), f, indent=2)
        total_size += metadata_file.stat().st_size
        
        items_exported['total_items'] = items_exported['assets'] + items_exported['threats']
        
        return ExportResult(
            success=True,
            output_path=output_path,
            format_used='csv_separate',
            file_size_bytes=total_size,
            export_time_ms=0,  # Will be updated by caller
            items_exported=items_exported,
            metadata=self._get_export_metadata(data),
            warnings=[]
        )
    
    def _export_single_csv_file(self, data: Dict[str, Any], 
                               output_path: Path, **options) -> ExportResult:
        """Export all data to a single CSV file with type column."""
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        # Combine assets and threats into single structure
        combined_data = []
        
        # Add assets
        for asset in data.get('assets', []):
            row = asset.copy()
            row['item_type'] = 'asset'
            # Flatten interfaces list
            if 'interfaces' in row and isinstance(row['interfaces'], list):
                row['interfaces'] = '; '.join(row['interfaces'])
            combined_data.append(row)
        
        # Add threats
        for threat in data.get('threats', []):
            row = threat.copy()
            row['item_type'] = 'threat'
            combined_data.append(row)
        
        if not combined_data:
            raise ExportError("No data to export")
        
        # Create DataFrame and export
        df = pd.DataFrame(combined_data)
        
        # Move item_type to first column
        columns = ['item_type'] + [col for col in df.columns if col != 'item_type']
        df = df[columns]
        
        csv_options = {
            'delimiter': options.get('delimiter', ','),
            'quoting': options.get('quoting', csv.QUOTE_MINIMAL),
            'lineterminator': '\n'
        }
        
        df.to_csv(output_path, index=False, **csv_options)
        
        return ExportResult(
            success=True,
            output_path=output_path,
            format_used='csv_combined',
            file_size_bytes=output_path.stat().st_size,
            export_time_ms=0,  # Will be updated by caller
            items_exported={
                'assets': len(data.get('assets', [])),
                'threats': len(data.get('threats', [])),
                'total_items': len(combined_data)
            },
            metadata=self._get_export_metadata(data),
            warnings=[]
        )
    
    def _reorder_columns(self, df: pd.DataFrame, preferred_order: List[str]) -> pd.DataFrame:
        """Reorder DataFrame columns with preferred order first."""
        existing_cols = df.columns.tolist()
        
        # Start with preferred columns that exist
        ordered_cols = [col for col in preferred_order if col in existing_cols]
        
        # Add remaining columns
        remaining_cols = [col for col in existing_cols if col not in ordered_cols]
        ordered_cols.extend(remaining_cols)
        
        return df[ordered_cols]


class ExcelExporter(BaseExporter):
    """Export TARA analysis results to Excel format."""
    
    def __init__(self):
        super().__init__('excel')
    
    def export(self, data: Dict[str, Any], output_path: Path, **options) -> ExportResult:
        """Export data to Excel format with multiple sheets.
        
        Args:
            data: Dictionary containing assets, threats, and metadata
            output_path: Path for output Excel file
            **options: Additional options (include_charts, formatting, etc.)
        
        Returns:
            ExportResult with export status and metadata
        """
        start_time = datetime.now()
        warnings = []
        
        try:
            # Validate input data
            validation_errors = self._validate_data(data)
            if validation_errors:
                raise ValidationError(f"Data validation failed: {'; '.join(validation_errors)}")
            
            output_path.parent.mkdir(parents=True, exist_ok=True)
            
            # Create Excel writer
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                items_exported = self._write_excel_sheets(writer, data, **options)
            
            # Calculate metrics
            processing_time = int((datetime.now() - start_time).total_seconds() * 1000)
            file_size = output_path.stat().st_size
            
            return ExportResult(
                success=True,
                output_path=output_path,
                format_used='excel',
                file_size_bytes=file_size,
                export_time_ms=processing_time,
                items_exported=items_exported,
                metadata=self._get_export_metadata(data),
                warnings=warnings
            )
            
        except Exception as e:
            logger.error(f"Excel export failed: {str(e)}")
            raise ExportError(f"Excel export failed: {str(e)}", format='excel')
    
    def _write_excel_sheets(self, writer: pd.ExcelWriter, 
                           data: Dict[str, Any], **options) -> Dict[str, int]:
        """Write data to Excel sheets."""
        items_exported = {'assets': 0, 'threats': 0}
        
        # Write assets sheet
        if 'assets' in data and data['assets']:
            assets_df = pd.DataFrame(data['assets'])
            
            # Process interfaces column for Excel
            if 'interfaces' in assets_df.columns:
                assets_df['interfaces'] = assets_df['interfaces'].apply(
                    lambda x: '; '.join(x) if isinstance(x, list) else str(x) if x else ''
                )
            
            # Reorder columns
            asset_columns = ['name', 'asset_type', 'criticality', 'interfaces', 
                           'description', 'location', 'manufacturer']
            assets_df = self._reorder_dataframe_columns(assets_df, asset_columns)
            
            assets_df.to_excel(writer, sheet_name='Assets', index=False)
            items_exported['assets'] = len(data['assets'])
            
            # Apply formatting if requested
            if options.get('apply_formatting', True):
                self._format_assets_sheet(writer, 'Assets', assets_df)
        
        # Write threats sheet
        if 'threats' in data and data['threats']:
            threats_df = pd.DataFrame(data['threats'])
            
            # Reorder columns
            threat_columns = ['name', 'category', 'description', 'likelihood', 
                            'impact', 'target_asset']
            threats_df = self._reorder_dataframe_columns(threats_df, threat_columns)
            
            threats_df.to_excel(writer, sheet_name='Threats', index=False)
            items_exported['threats'] = len(data['threats'])
            
            if options.get('apply_formatting', True):
                self._format_threats_sheet(writer, 'Threats', threats_df)
        
        # Write summary sheet
        if options.get('include_summary', True):
            self._write_summary_sheet(writer, data, items_exported)
        
        # Write metadata sheet
        self._write_metadata_sheet(writer, data)
        
        items_exported['total_items'] = items_exported['assets'] + items_exported['threats']
        return items_exported
    
    def _reorder_dataframe_columns(self, df: pd.DataFrame, 
                                  preferred_order: List[str]) -> pd.DataFrame:
        """Reorder DataFrame columns with preferred order first."""
        existing_cols = df.columns.tolist()
        
        # Start with preferred columns that exist
        ordered_cols = [col for col in preferred_order if col in existing_cols]
        
        # Add remaining columns
        remaining_cols = [col for col in existing_cols if col not in ordered_cols]
        ordered_cols.extend(remaining_cols)
        
        return df[ordered_cols]
    
    def _format_assets_sheet(self, writer: pd.ExcelWriter, sheet_name: str, df: pd.DataFrame):
        """Apply formatting to assets sheet."""
        try:
            workbook = writer.book
            worksheet = writer.sheets[sheet_name]
            
            # Auto-adjust column widths
            for column in df:
                column_length = max(df[column].astype(str).map(len).max(), len(column))
                col_idx = df.columns.get_loc(column)
                worksheet.column_dimensions[chr(65 + col_idx)].width = min(column_length + 2, 50)
            
            # Add header formatting (if openpyxl available)
            if hasattr(worksheet, 'cell'):
                from openpyxl.styles import Font, PatternFill
                
                header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                header_font = Font(color='FFFFFF', bold=True)
                
                for col_num, column_title in enumerate(df.columns, 1):
                    cell = worksheet.cell(row=1, column=col_num)
                    cell.fill = header_fill
                    cell.font = header_font
                    
        except ImportError:
            # openpyxl not available, skip formatting
            pass
        except Exception as e:
            logger.warning(f"Could not apply Excel formatting: {str(e)}")
    
    def _format_threats_sheet(self, writer: pd.ExcelWriter, sheet_name: str, df: pd.DataFrame):
        """Apply formatting to threats sheet."""
        # Similar to assets formatting but with threat-specific styling
        self._format_assets_sheet(writer, sheet_name, df)
    
    def _write_summary_sheet(self, writer: pd.ExcelWriter, data: Dict[str, Any], 
                            items_exported: Dict[str, int]):
        """Write summary information sheet."""
        summary_data = [
            ['Analysis Summary', ''],
            ['Total Assets', items_exported['assets']],
            ['Total Threats', items_exported['threats']],
            ['Export Date', self.export_timestamp.strftime('%Y-%m-%d %H:%M:%S')],
            ['', ''],
        ]
        
        # Add criticality breakdown if available
        if 'assets' in data and data['assets']:
            criticality_counts = {}
            for asset in data['assets']:
                crit = asset.get('criticality', 'Unknown')
                criticality_counts[crit] = criticality_counts.get(crit, 0) + 1
            
            summary_data.append(['Asset Criticality Breakdown', ''])
            for crit, count in sorted(criticality_counts.items()):
                summary_data.append([f'  {crit}', count])
        
        summary_df = pd.DataFrame(summary_data, columns=['Metric', 'Value'])
        summary_df.to_excel(writer, sheet_name='Summary', index=False)
    
    def _write_metadata_sheet(self, writer: pd.ExcelWriter, data: Dict[str, Any]):
        """Write metadata information sheet."""
        metadata = self._get_export_metadata(data)
        
        metadata_rows = []
        for key, value in metadata.items():
            if isinstance(value, (dict, list)):
                value = json.dumps(value, indent=2)
            metadata_rows.append([key, str(value)])
        
        metadata_df = pd.DataFrame(metadata_rows, columns=['Property', 'Value'])
        metadata_df.to_excel(writer, sheet_name='Metadata', index=False)


class ExportManager:
    """Manages export operations for different formats."""
    
    def __init__(self):
        """Initialize export manager with available exporters."""
        self.exporters = {
            'json': JsonExporter(),
            'csv': CsvExporter(),
            'excel': ExcelExporter()
        }
    
    def export_data(self, data: Dict[str, Any], output_path: Union[str, Path],
                   format_name: str = None, **options) -> ExportResult:
        """Export data to specified format and path.
        
        Args:
            data: Dictionary containing assets, threats, and metadata
            output_path: Output file or directory path
            format_name: Export format ('json', 'csv', 'excel') - auto-detected if None
            **options: Format-specific options
        
        Returns:
            ExportResult with export status and metadata
        """
        output_path = Path(output_path)
        
        # Auto-detect format if not specified
        if format_name is None:
            format_name = self._detect_format_from_path(output_path)
        
        if format_name not in self.exporters:
            raise ExportError(f"Unsupported export format: {format_name}")
        
        exporter = self.exporters[format_name]
        return exporter.export(data, output_path, **options)
    
    def get_supported_formats(self) -> List[str]:
        """Get list of supported export formats."""
        return list(self.exporters.keys())
    
    def _detect_format_from_path(self, path: Path) -> str:
        """Detect export format from file extension."""
        extension = path.suffix.lower()
        
        mapping = {
            '.json': 'json',
            '.csv': 'csv',
            '.xlsx': 'excel',
            '.xls': 'excel'
        }
        
        format_name = mapping.get(extension)
        if not format_name:
            raise ExportError(f"Cannot determine format from extension: {extension}")
        
        return format_name


# Utility functions
def export_tara_data(data: Dict[str, Any], output_path: Union[str, Path],
                    format_name: str = None, **options) -> ExportResult:
    """Convenience function to export TARA data."""
    manager = ExportManager()
    return manager.export_data(data, output_path, format_name, **options)


def get_export_formats() -> List[str]:
    """Get list of supported export formats."""
    manager = ExportManager()
    return manager.get_supported_formats()