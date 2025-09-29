"""File handler for multi-format input processing.

Reference: plan.md lines 75-76 (multi-format support)
Supports Excel (.xlsx), CSV, JSON, text files with 10MB size limit validation.
"""

import os
import json
import mimetypes
from pathlib import Path
from typing import Dict, Any, List, Optional, Union
from dataclasses import dataclass
from io import BytesIO
import pandas as pd
from openpyxl import load_workbook


@dataclass
class FileValidationResult:
    """Result of file validation."""
    is_valid: bool
    file_size_mb: float
    detected_format: Optional[str]
    error_message: Optional[str] = None


@dataclass 
class ParsedFileData:
    """Parsed file data with metadata."""
    data: Union[Dict[str, Any], List[Dict[str, Any]]]
    format: str
    size_mb: float
    columns: Optional[List[str]] = None
    row_count: Optional[int] = None


class FileHandlerError(Exception):
    """Custom exception for file handling operations."""
    pass


class FileHandler:
    """Multi-format file handler for TARA input processing.
    
    Reference: plan.md multi-format support requirements
    """
    
    # 10MB file size limit as specified in requirements
    MAX_FILE_SIZE_MB = 10.0
    MAX_FILE_SIZE_BYTES = MAX_FILE_SIZE_MB * 1024 * 1024
    
    # Supported formats and extensions
    SUPPORTED_FORMATS = {
        'excel': ['.xlsx', '.xls'],
        'csv': ['.csv'],
        'json': ['.json'],
        'text': ['.txt', '.md']
    }
    
    def __init__(self):
        """Initialize file handler."""
        self._setup_mime_types()
    
    def _setup_mime_types(self) -> None:
        """Setup additional MIME type mappings."""
        # Add custom MIME types for better detection
        mimetypes.add_type('application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', '.xlsx')
        mimetypes.add_type('application/vnd.ms-excel', '.xls')
    
    def validate_file(self, file_path: Union[str, Path]) -> FileValidationResult:
        """Validate file exists, format, and size requirements.
        
        Args:
            file_path: Path to file to validate
            
        Returns:
            FileValidationResult with validation details
        """
        file_path = Path(file_path)
        
        # Check file exists
        if not file_path.exists():
            return FileValidationResult(
                is_valid=False,
                file_size_mb=0.0,
                detected_format=None,
                error_message=f"File does not exist: {file_path}"
            )
        
        # Check file size
        file_size_bytes = file_path.stat().st_size
        file_size_mb = file_size_bytes / (1024 * 1024)
        
        if file_size_bytes > self.MAX_FILE_SIZE_BYTES:
            return FileValidationResult(
                is_valid=False,
                file_size_mb=file_size_mb,
                detected_format=None,
                error_message=f"File too large: {file_size_mb:.2f}MB exceeds {self.MAX_FILE_SIZE_MB}MB limit"
            )
        
        # Detect format
        detected_format = self._detect_format(file_path)
        if not detected_format:
            return FileValidationResult(
                is_valid=False,
                file_size_mb=file_size_mb,
                detected_format=None,
                error_message=f"Unsupported file format: {file_path.suffix}"
            )
        
        return FileValidationResult(
            is_valid=True,
            file_size_mb=file_size_mb,
            detected_format=detected_format
        )
    
    def _detect_format(self, file_path: Path) -> Optional[str]:
        """Detect file format from extension and MIME type.
        
        Args:
            file_path: Path to file
            
        Returns:
            Detected format string or None if unsupported
        """
        extension = file_path.suffix.lower()
        
        # Check against supported formats
        for format_name, extensions in self.SUPPORTED_FORMATS.items():
            if extension in extensions:
                return format_name
        
        return None
    
    def parse_file(self, file_path: Union[str, Path]) -> ParsedFileData:
        """Parse file and return structured data.
        
        Args:
            file_path: Path to file to parse
            
        Returns:
            ParsedFileData with parsed content and metadata
        """
        file_path = Path(file_path)
        
        # Validate file first
        validation = self.validate_file(file_path)
        if not validation.is_valid:
            raise FileHandlerError(f"File validation failed: {validation.error_message}")
        
        # Parse based on detected format
        format_type = validation.detected_format
        
        try:
            if format_type == 'excel':
                return self._parse_excel(file_path, validation.file_size_mb)
            elif format_type == 'csv':
                return self._parse_csv(file_path, validation.file_size_mb)
            elif format_type == 'json':
                return self._parse_json(file_path, validation.file_size_mb)
            elif format_type == 'text':
                return self._parse_text(file_path, validation.file_size_mb)
            else:
                raise FileHandlerError(f"Unsupported format: {format_type}")
        
        except Exception as e:
            raise FileHandlerError(f"Failed to parse {format_type} file: {e}")
    
    def _parse_excel(self, file_path: Path, size_mb: float) -> ParsedFileData:
        """Parse Excel file using pandas and openpyxl."""
        try:
            # Use pandas for data extraction
            df = pd.read_excel(file_path, engine='openpyxl')
            
            # Convert to list of dictionaries
            data = df.to_dict('records')
            
            return ParsedFileData(
                data=data,
                format='excel',
                size_mb=size_mb,
                columns=list(df.columns),
                row_count=len(df)
            )
        
        except Exception as e:
            # Try with openpyxl directly for more control
            try:
                workbook = load_workbook(file_path, data_only=True)
                sheet = workbook.active
                
                # Get headers from first row
                headers = [cell.value for cell in sheet[1]]
                
                # Get data rows
                data = []
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    row_dict = dict(zip(headers, row))
                    data.append(row_dict)
                
                return ParsedFileData(
                    data=data,
                    format='excel',
                    size_mb=size_mb,
                    columns=headers,
                    row_count=len(data)
                )
            
            except Exception as inner_e:
                raise FileHandlerError(f"Excel parsing failed: {e}, {inner_e}")
    
    def _parse_csv(self, file_path: Path, size_mb: float) -> ParsedFileData:
        """Parse CSV file using pandas."""
        try:
            # Try different encodings
            for encoding in ['utf-8', 'utf-8-sig', 'latin-1', 'cp1252']:
                try:
                    df = pd.read_csv(file_path, encoding=encoding)
                    break
                except UnicodeDecodeError:
                    continue
            else:
                raise FileHandlerError("Could not decode CSV file with any supported encoding")
            
            # Convert to list of dictionaries
            data = df.to_dict('records')
            
            return ParsedFileData(
                data=data,
                format='csv',
                size_mb=size_mb,
                columns=list(df.columns),
                row_count=len(df)
            )
        
        except Exception as e:
            raise FileHandlerError(f"CSV parsing failed: {e}")
    
    def _parse_json(self, file_path: Path, size_mb: float) -> ParsedFileData:
        """Parse JSON file."""
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            # Ensure data is in list format for consistency
            if isinstance(data, dict):
                data = [data]
            elif not isinstance(data, list):
                raise FileHandlerError("JSON must contain an object or array of objects")
            
            # Extract columns if data is list of dictionaries
            columns = None
            if data and isinstance(data[0], dict):
                columns = list(data[0].keys())
            
            return ParsedFileData(
                data=data,
                format='json',
                size_mb=size_mb,
                columns=columns,
                row_count=len(data) if isinstance(data, list) else 1
            )
        
        except json.JSONDecodeError as e:
            raise FileHandlerError(f"Invalid JSON format: {e}")
        except Exception as e:
            raise FileHandlerError(f"JSON parsing failed: {e}")
    
    def _parse_text(self, file_path: Path, size_mb: float) -> ParsedFileData:
        """Parse text file."""
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                content = f.read()
            
            # Split into lines and create simple structure
            lines = [line.strip() for line in content.split('\n') if line.strip()]
            
            # Create list of dictionaries with line numbers
            data = [
                {'line_number': i + 1, 'content': line}
                for i, line in enumerate(lines)
            ]
            
            return ParsedFileData(
                data=data,
                format='text',
                size_mb=size_mb,
                columns=['line_number', 'content'],
                row_count=len(data)
            )
        
        except Exception as e:
            raise FileHandlerError(f"Text parsing failed: {e}")
    
    def validate_file_from_bytes(self, file_bytes: bytes, filename: str) -> FileValidationResult:
        """Validate file from bytes data (for upload scenarios).
        
        Args:
            file_bytes: File content as bytes
            filename: Original filename for format detection
            
        Returns:
            FileValidationResult with validation details
        """
        file_size_bytes = len(file_bytes)
        file_size_mb = file_size_bytes / (1024 * 1024)
        
        # Check file size
        if file_size_bytes > self.MAX_FILE_SIZE_BYTES:
            return FileValidationResult(
                is_valid=False,
                file_size_mb=file_size_mb,
                detected_format=None,
                error_message=f"File too large: {file_size_mb:.2f}MB exceeds {self.MAX_FILE_SIZE_MB}MB limit"
            )
        
        # Detect format from filename
        file_path = Path(filename)
        detected_format = self._detect_format(file_path)
        if not detected_format:
            return FileValidationResult(
                is_valid=False,
                file_size_mb=file_size_mb,
                detected_format=None,
                error_message=f"Unsupported file format: {file_path.suffix}"
            )
        
        return FileValidationResult(
            is_valid=True,
            file_size_mb=file_size_mb,
            detected_format=detected_format
        )
    
    def get_supported_extensions(self) -> List[str]:
        """Get list of all supported file extensions."""
        extensions = []
        for format_extensions in self.SUPPORTED_FORMATS.values():
            extensions.extend(format_extensions)
        return sorted(extensions)