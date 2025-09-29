"""Export service for JSON/Excel generation.

Reference: contracts/api.yaml lines 147-190 (export endpoint)
Generates structured JSON with ISO compliance fields and Excel spreadsheets.
"""

import json
import os
from datetime import datetime
from pathlib import Path
from typing import Dict, Any, List, Optional, Union
from dataclasses import dataclass
from io import BytesIO
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

from ..models import TaraAnalysis
from .database import DatabaseService


@dataclass
class ExportConfig:
    """Configuration for export operations."""
    include_metadata: bool = True
    include_timestamps: bool = True
    include_iso_sections: bool = True
    excel_styling: bool = True
    output_directory: Optional[str] = None


@dataclass
class ExportResult:
    """Result of export operation."""
    success: bool
    file_path: Optional[str] = None
    file_size_bytes: Optional[int] = None
    format: Optional[str] = None
    error_message: Optional[str] = None


class ExportError(Exception):
    """Custom exception for export operations."""
    pass


class ExportService:
    """Export service for generating JSON and Excel outputs.
    
    Reference: contracts/api.yaml export format requirements
    """
    
    def __init__(self, db_service: DatabaseService, config: Optional[ExportConfig] = None):
        """Initialize export service.
        
        Args:
            db_service: Database service instance
            config: Export configuration
        """
        self.db_service = db_service
        self.config = config or ExportConfig()
        self._setup_output_directory()
    
    def _setup_output_directory(self) -> None:
        """Setup output directory for exports."""
        if self.config.output_directory:
            output_dir = Path(self.config.output_directory)
        else:
            output_dir = Path.cwd() / "exports"
        
        output_dir.mkdir(exist_ok=True)
        self.output_directory = output_dir
    
    def export_analysis_json(self, analysis_id: str, file_path: Optional[str] = None) -> ExportResult:
        """Export TARA analysis to structured JSON format.
        
        Args:
            analysis_id: ID of analysis to export
            file_path: Optional output file path
            
        Returns:
            ExportResult with export details
        """
        try:
            # Generate JSON data
            json_data = self._generate_analysis_json(analysis_id)
            
            # Determine output path
            if not file_path:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                file_path = self.output_directory / f"tara_analysis_{analysis_id[:8]}_{timestamp}.json"
            
            file_path = Path(file_path)
            
            # Write JSON file
            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(json_data, f, indent=2, ensure_ascii=False, default=str)
            
            file_size = file_path.stat().st_size
            
            return ExportResult(
                success=True,
                file_path=str(file_path),
                file_size_bytes=file_size,
                format='json'
            )
            
        except Exception as e:
            return ExportResult(
                success=False,
                error_message=f"JSON export failed: {e}"
            )
    
    def export_analysis_excel(self, analysis_id: str, file_path: Optional[str] = None) -> ExportResult:
        """Export TARA analysis to Excel spreadsheet format.
        
        Args:
            analysis_id: ID of analysis to export
            file_path: Optional output file path
            
        Returns:
            ExportResult with export details
        """
        try:
            # Generate Excel data
            workbook = self._generate_analysis_excel(analysis_id)
            
            # Determine output path
            if not file_path:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                file_path = self.output_directory / f"tara_analysis_{analysis_id[:8]}_{timestamp}.xlsx"
            
            file_path = Path(file_path)
            
            # Save workbook
            workbook.save(file_path)
            
            file_size = file_path.stat().st_size
            
            return ExportResult(
                success=True,
                file_path=str(file_path),
                file_size_bytes=file_size,
                format='excel'
            )
            
        except Exception as e:
            return ExportResult(
                success=False,
                error_message=f"Excel export failed: {e}"
            )
    
    def _generate_analysis_json(self, analysis_id: str) -> Dict[str, Any]:
        """Generate structured JSON data for analysis.
        
        Args:
            analysis_id: Analysis ID (can be full UUID or partial)
            
        Returns:
            Dictionary with structured analysis data
        """
        with self.db_service.get_session() as session:
            # Load analysis with all relationships
            from ..models import TaraAnalysis
            from sqlalchemy.orm import selectinload
            from sqlalchemy import func
            
            # Convert partial ID to full UUID if needed
            from uuid import UUID
            try:
                # Try to parse as full UUID
                uuid_obj = UUID(analysis_id)
                full_analysis_id = str(uuid_obj)
            except ValueError:
                # If partial ID, find the matching full ID
                partial_matches = session.query(TaraAnalysis).filter(
                    func.cast(TaraAnalysis.id, str).like(f"{analysis_id}%")
                ).all()
                
                if len(partial_matches) == 0:
                    raise ExportError(f"No analysis found matching ID: {analysis_id}")
                elif len(partial_matches) > 1:
                    matching_ids = [str(a.id)[:8] for a in partial_matches]
                    raise ExportError(f"Multiple analyses found matching '{analysis_id}': {matching_ids}. Please use a more specific ID.")
                else:
                    full_analysis_id = str(partial_matches[0].id)
            
            # Query just the basic analysis first
            analysis = session.query(TaraAnalysis).filter(TaraAnalysis.id == full_analysis_id).first()
            
            if not analysis:
                raise ExportError(f"Analysis not found: {analysis_id}")
            
            # Count related entities separately to avoid relationship loading issues
            from ..models import Asset, CybersecurityGoal
            asset_count = session.query(Asset).filter(Asset.analysis_id == full_analysis_id).count()
            goal_count = session.query(CybersecurityGoal).filter(CybersecurityGoal.analysis_id == full_analysis_id).count()
            
            # Build basic JSON structure without complex relationships
            json_data = {
                "analysis_metadata": {
                    "id": str(analysis.id),
                    "analysis_name": analysis.analysis_name,
                    "vehicle_model": analysis.vehicle_model,
                    "analysis_phase": analysis.analysis_phase.value,
                    "completion_status": analysis.completion_status.value,
                    "input_file_path": analysis.input_file_path,
                    "output_file_path": analysis.output_file_path,
                    "created_at": analysis.created_at.isoformat() if self.config.include_timestamps and analysis.created_at else None,
                    "completed_at": analysis.completed_at.isoformat() if analysis.completed_at and self.config.include_timestamps else None,
                    "iso_section": analysis.iso_section if self.config.include_iso_sections else None,
                },
                "statistics": {
                    "total_assets": asset_count,
                    "total_goals": goal_count,
                },
                "assets": [],
                "cybersecurity_goals": [],
                "export_metadata": {
                    "exported_at": datetime.now().isoformat() if self.config.include_metadata else None,
                    "export_version": "1.0",
                    "iso_compliance": "ISO/SAE 21434",
                    "note": "Basic export - detailed relationships available in full export mode"
                }
            }
            
            # Try to load assets without complex relationships
            try:
                assets = session.query(Asset).filter(Asset.analysis_id == full_analysis_id).all()
                for asset in assets:
                    asset_data = {
                        "id": str(asset.id),
                        "name": asset.name,
                        "asset_type": asset.asset_type.value if asset.asset_type else None,
                        "criticality_level": asset.criticality_level.value if asset.criticality_level else None
                    }
                    json_data["assets"].append(asset_data)
            except Exception as e:
                json_data["assets"] = [{"error": f"Could not load assets: {e}"}]
            
            # Try to load cybersecurity goals without complex relationships  
            try:
                goals = session.query(CybersecurityGoal).filter(CybersecurityGoal.analysis_id == full_analysis_id).all()
                for goal in goals:
                    goal_data = {
                        "id": str(goal.id),
                        "goal_name": goal.goal_name,
                        "protection_level": goal.protection_level.value if goal.protection_level else None,
                        "implementation_phase": goal.implementation_phase.value if goal.implementation_phase else None
                    }
                    json_data["cybersecurity_goals"].append(goal_data)
            except Exception as e:
                json_data["cybersecurity_goals"] = [{"error": f"Could not load goals: {e}"}]
            
            return json_data
    
    def _generate_analysis_excel(self, analysis_id: str) -> Workbook:
        """Generate Excel workbook for analysis.
        
        Args:
            analysis_id: Analysis ID
            
        Returns:
            Openpyxl Workbook instance
        """
        # Get JSON data first
        json_data = self._generate_analysis_json(analysis_id)
        
        # Create workbook
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create sheets for different data types
        self._create_summary_sheet(wb, json_data)
        self._create_assets_sheet(wb, json_data)
        self._create_threats_sheet(wb, json_data)
        self._create_risks_sheet(wb, json_data)
        self._create_treatments_sheet(wb, json_data)
        self._create_goals_sheet(wb, json_data)
        
        return wb
    
    def _create_summary_sheet(self, wb: Workbook, data: Dict[str, Any]) -> None:
        """Create summary sheet in Excel workbook."""
        ws = wb.create_sheet("Summary", 0)
        
        # Headers
        headers = ["Property", "Value"]
        ws.append(headers)
        
        # Analysis metadata
        metadata = data["analysis_metadata"]
        for key, value in metadata.items():
            if value is not None:
                ws.append([key.replace('_', ' ').title(), str(value)])
        
        # Statistics
        ws.append(["", ""])  # Empty row
        ws.append(["Statistics", ""])
        ws.append(["Total Assets", len(data["assets"])])
        ws.append(["Total Goals", len(data["cybersecurity_goals"])])
        
        if self.config.excel_styling:
            self._apply_header_styling(ws, headers)
    
    def _create_assets_sheet(self, wb: Workbook, data: Dict[str, Any]) -> None:
        """Create assets sheet in Excel workbook."""
        if not data["assets"]:
            return
        
        ws = wb.create_sheet("Assets")
        
        # Headers
        headers = [
            "Asset ID", "Name", "Type", "Criticality", "Interfaces", 
            "Data Flows", "Security Properties", "ISO Section"
        ]
        ws.append(headers)
        
        # Asset data
        for asset in data["assets"]:
            row = [
                asset.get("id", ""),
                asset.get("name", ""),
                asset.get("asset_type", ""),
                asset.get("criticality_level", ""),
                ", ".join(asset.get("interfaces", [])),
                ", ".join(asset.get("data_flows", [])),
                json.dumps(asset.get("security_properties", {})),
                asset.get("iso_section", "")
            ]
            ws.append(row)
        
        if self.config.excel_styling:
            self._apply_header_styling(ws, headers)
    
    def _create_threats_sheet(self, wb: Workbook, data: Dict[str, Any]) -> None:
        """Create threats sheet in Excel workbook."""
        ws = wb.create_sheet("Threats")
        
        headers = [
            "Threat ID", "Asset Name", "Threat Name", "Actor", "Motivation",
            "Attack Vectors", "Prerequisites", "ISO Section"
        ]
        ws.append(headers)
        
        # Threat scenarios from assets
        for asset in data["assets"]:
            for threat in asset.get("threat_scenarios", []):
                row = [
                    threat.get("id", ""),
                    asset.get("name", ""),
                    threat.get("threat_name", ""),
                    threat.get("threat_actor", ""),
                    threat.get("motivation", ""),
                    ", ".join(threat.get("attack_vectors", [])),
                    ", ".join(threat.get("prerequisites", [])),
                    threat.get("iso_section", "")
                ]
                ws.append(row)
        
        if self.config.excel_styling:
            self._apply_header_styling(ws, headers)
    
    def _create_risks_sheet(self, wb: Workbook, data: Dict[str, Any]) -> None:
        """Create risks sheet in Excel workbook."""
        ws = wb.create_sheet("Risks")
        
        headers = [
            "Risk ID", "Asset Name", "Threat Name", "Risk Level", "Risk Score",
            "Impact Score", "Feasibility Score", "Calculation Method"
        ]
        ws.append(headers)
        
        # Risk values from assets
        for asset in data["assets"]:
            for risk in asset.get("risk_values", []):
                row = [
                    risk.get("id", ""),
                    asset.get("name", ""),
                    risk.get("threat_scenario_name", ""),
                    risk.get("risk_level", ""),
                    risk.get("risk_score", ""),
                    risk.get("impact_score", ""),
                    risk.get("feasibility_score", ""),
                    risk.get("calculation_method", "")
                ]
                ws.append(row)
        
        if self.config.excel_styling:
            self._apply_header_styling(ws, headers)
    
    def _create_treatments_sheet(self, wb: Workbook, data: Dict[str, Any]) -> None:
        """Create treatments sheet in Excel workbook."""
        ws = wb.create_sheet("Treatments")
        
        headers = [
            "Treatment ID", "Risk ID", "Decision", "Countermeasures",
            "Residual Risk", "Cost", "Rationale", "ISO Section"
        ]
        ws.append(headers)
        
        # Risk treatments
        for asset in data["assets"]:
            for risk in asset.get("risk_values", []):
                treatment = risk.get("risk_treatment")
                if treatment:
                    row = [
                        treatment.get("id", ""),
                        risk.get("id", ""),
                        treatment.get("treatment_decision", ""),
                        ", ".join(treatment.get("countermeasures", [])),
                        treatment.get("residual_risk_level", ""),
                        treatment.get("implementation_cost", ""),
                        treatment.get("rationale", ""),
                        treatment.get("iso_section", "")
                    ]
                    ws.append(row)
        
        if self.config.excel_styling:
            self._apply_header_styling(ws, headers)
    
    def _create_goals_sheet(self, wb: Workbook, data: Dict[str, Any]) -> None:
        """Create cybersecurity goals sheet in Excel workbook."""
        ws = wb.create_sheet("Goals")
        
        headers = [
            "Goal ID", "Goal Name", "Protection Level", "Security Controls",
            "Verification Method", "Implementation Phase", "ISO Section"
        ]
        ws.append(headers)
        
        # Cybersecurity goals
        for goal in data["cybersecurity_goals"]:
            row = [
                goal.get("id", ""),
                goal.get("goal_name", ""),
                goal.get("protection_level", ""),
                ", ".join(goal.get("security_controls", [])),
                goal.get("verification_method", ""),
                goal.get("implementation_phase", ""),
                goal.get("iso_section", "")
            ]
            ws.append(row)
        
        if self.config.excel_styling:
            self._apply_header_styling(ws, headers)
    
    def _apply_header_styling(self, worksheet, headers: List[str]) -> None:
        """Apply styling to worksheet headers."""
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
    
    def _serialize_asset(self, asset) -> Dict[str, Any]:
        """Serialize asset model to dictionary."""
        return {
            "id": str(asset.id),
            "name": asset.name,
            "asset_type": asset.asset_type.value,
            "criticality_level": asset.criticality_level.value,
            "interfaces": asset.interfaces,
            "data_flows": asset.data_flows,
            "security_properties": asset.security_properties,
            "iso_section": asset.iso_section if self.config.include_iso_sections else None,
            "created_at": asset.created_at.isoformat() if self.config.include_timestamps else None,
            "threat_scenarios": [self._serialize_threat_scenario(ts) for ts in asset.threat_scenarios],
            "impact_ratings": [self._serialize_impact_rating(ir) for ir in asset.impact_ratings],
            "risk_values": [self._serialize_risk_value(rv) for rv in asset.risk_values],
        }
    
    def _serialize_threat_scenario(self, threat) -> Dict[str, Any]:
        """Serialize threat scenario model to dictionary."""
        return {
            "id": str(threat.id),
            "threat_name": threat.threat_name,
            "threat_actor": threat.threat_actor.value,
            "motivation": threat.motivation,
            "attack_vectors": threat.attack_vectors,
            "prerequisites": threat.prerequisites,
            "iso_section": threat.iso_section if self.config.include_iso_sections else None,
            "attack_paths": [self._serialize_attack_path(ap) for ap in threat.attack_paths]
        }
    
    def _serialize_attack_path(self, path) -> Dict[str, Any]:
        """Serialize attack path model to dictionary."""
        return {
            "id": str(path.id),
            "step_sequence": path.step_sequence,
            "attack_step": path.attack_step,
            "intermediate_targets": path.intermediate_targets,
            "technical_barriers": path.technical_barriers,
            "required_resources": path.required_resources,
            "attack_feasibility": self._serialize_attack_feasibility(path.attack_feasibility) if path.attack_feasibility else None
        }
    
    def _serialize_attack_feasibility(self, feasibility) -> Dict[str, Any]:
        """Serialize attack feasibility model to dictionary."""
        return {
            "id": str(feasibility.id),
            "elapsed_time": feasibility.elapsed_time.value,
            "specialist_expertise": feasibility.specialist_expertise.value,
            "knowledge_of_target": feasibility.knowledge_of_target.value,
            "window_of_opportunity": feasibility.window_of_opportunity.value,
            "equipment_required": feasibility.equipment_required.value,
            "feasibility_score": feasibility.feasibility_score
        }
    
    def _serialize_impact_rating(self, impact) -> Dict[str, Any]:
        """Serialize impact rating model to dictionary."""
        return {
            "id": str(impact.id),
            "safety_impact": impact.safety_impact.value,
            "financial_impact": impact.financial_impact.value,
            "operational_impact": impact.operational_impact.value,
            "privacy_impact": impact.privacy_impact.value,
            "impact_score": impact.impact_score,
            "iso_section": impact.iso_section if self.config.include_iso_sections else None
        }
    
    def _serialize_risk_value(self, risk) -> Dict[str, Any]:
        """Serialize risk value model to dictionary."""
        return {
            "id": str(risk.id),
            "risk_level": risk.risk_level.value,
            "risk_score": risk.risk_score,
            "calculation_method": risk.calculation_method,
            "impact_score": risk.impact_rating.impact_score if risk.impact_rating else None,
            "feasibility_score": risk.attack_feasibility.feasibility_score if risk.attack_feasibility else None,
            "threat_scenario_name": risk.threat_scenario.threat_name if risk.threat_scenario else None,
            "risk_treatment": self._serialize_risk_treatment(risk.risk_treatment) if risk.risk_treatment else None
        }
    
    def _serialize_risk_treatment(self, treatment) -> Dict[str, Any]:
        """Serialize risk treatment model to dictionary."""
        return {
            "id": str(treatment.id),
            "treatment_decision": treatment.treatment_decision.value,
            "countermeasures": treatment.countermeasures,
            "residual_risk_level": treatment.residual_risk_level.value,
            "implementation_cost": treatment.implementation_cost,
            "rationale": treatment.rationale,
            "iso_section": treatment.iso_section if self.config.include_iso_sections else None
        }
    
    def _serialize_cybersecurity_goal(self, goal) -> Dict[str, Any]:
        """Serialize cybersecurity goal model to dictionary."""
        return {
            "id": str(goal.id),
            "goal_name": goal.goal_name,
            "protection_level": goal.protection_level.value,
            "security_controls": goal.security_controls,
            "verification_method": goal.verification_method,
            "implementation_phase": goal.implementation_phase.value,
            "iso_section": goal.iso_section if self.config.include_iso_sections else None
        }
    
    def get_export_formats(self) -> List[str]:
        """Get list of supported export formats."""
        return ["json", "excel"]
    
    def cleanup_old_exports(self, days_old: int = 30) -> int:
        """Clean up old export files.
        
        Args:
            days_old: Files older than this many days will be deleted
            
        Returns:
            Number of files deleted
        """
        import time
        
        cutoff_time = time.time() - (days_old * 24 * 60 * 60)
        deleted_count = 0
        
        for file_path in self.output_directory.iterdir():
            if file_path.is_file() and file_path.stat().st_mtime < cutoff_time:
                try:
                    file_path.unlink()
                    deleted_count += 1
                except Exception:
                    pass  # Ignore errors for individual file deletions
        
        return deleted_count